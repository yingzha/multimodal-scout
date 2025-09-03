"""
FastAPI backend server for Multimodal Scout application.
Provides REST API endpoints for fetching topics and scraping content.
"""

from alembic.config import Config
from alembic import command
from fastapi import FastAPI, HTTPException, Header, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, Response
from typing import Optional
from contextlib import asynccontextmanager
import asyncio
import json
import time
from functools import lru_cache
from sqlalchemy import text

from .constants import INTERESTED_KEYWORDS
from .logger import logger
from .database import db_manager
from .pipeline import process_content_pipeline
from .schema import (
    FetchRequest,
    TopicResponse,
    ItemResponse,
    FetchResponse,
    BookmarkRequest,
    BookmarkResponse,
    UploadLinkRequest,
    UploadLinkResponse,
    UserRegistrationRequest,
    UserLoginRequest,
    AuthResponse,
    UserResponse,
    ConfigResponse,
    UserPreferencesResponse,
    UpdateUserPreferencesRequest,
)
from .utils import (
    generate_summary_from_link,
    extract_title_from_url,
    categorize_content,
    _fetch_article_text,
)


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Handle application startup and shutdown"""
    # Startup
    try:
        logger.info("🚀 Initializing database tables...")
        db_manager.create_tables()

        # Check if we need to run migrations automatically
        try:
            logger.info(
                "🔄 Checking database schema and running migrations if needed..."
            )

            # Test database connection first

            try:
                with db_manager.engine.connect() as conn:
                    conn.execute(text("SELECT 1"))
                    logger.info("✅ Database connection verified")
            except Exception as conn_error:
                logger.warning(
                    f"⚠️ Database connection failed: {conn_error} - migration skipped"
                )
                logger.warning(
                    "💡 Manual migration may be needed using Cloud SQL Proxy"
                )
                return

            # Set up alembic configuration
            alembic_cfg = Config("/app/alembic.ini")
            alembic_cfg.set_main_option("script_location", "/app/alembic")

            # Run migration
            logger.info("🔄 Running database migrations...")
            command.upgrade(alembic_cfg, "head")
            logger.info("✅ Database migrations completed successfully")

        except Exception as migration_error:
            logger.error(
                f"❌ Failed to run migrations: {migration_error}", exc_info=True
            )
            logger.warning(
                "⚠️ Migration failed - manual intervention may be needed. Service will continue to start."
            )
            logger.warning(
                "💡 Use Cloud SQL Proxy method for manual migration if needed"
            )
            # Don't fail startup if migrations fail - allow manual intervention

        logger.info("✅ Database initialization completed")
    except Exception as e:
        logger.error(f"❌ Failed to initialize database: {e}", exc_info=True)
        # Don't raise here to allow the app to start even if DB init fails
        # This allows for debugging and manual intervention

    yield

    # Shutdown (if needed)
    logger.info("🔄 Application shutting down...")


app = FastAPI(
    title="Multimodal Scout API",
    description="API for fetching AI/ML content from various sources",
    version="1.0.0",
    lifespan=lifespan,
)


# Helper function for user-friendly error responses
def create_user_friendly_error(
    error_type: str,
    user_message: str,
    technical_detail: str = None,
    status_code: int = 500,
):
    """Create a user-friendly error response"""
    logger.error(f"{error_type}: {technical_detail or user_message}")

    # Map common error types to user-friendly messages
    error_messages = {
        "database_error": "We're having trouble accessing our database. Please try again in a moment.",
        "external_api_error": "We're having trouble connecting to external services. Please try again.",
        "validation_error": "The information provided doesn't meet our requirements. Please check and try again.",
        "not_found_error": "The requested item could not be found.",
        "rate_limit_error": "Too many requests. Please wait a moment before trying again.",
        "processing_error": "We're having trouble processing your request. Please try again.",
    }

    final_message = error_messages.get(error_type, user_message)

    raise HTTPException(
        status_code=status_code,
        detail={
            "error": error_type,
            "message": final_message,
            "details": technical_detail if technical_detail else None,
        },
    )


# Add CORS middleware to allow frontend connections
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for development
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# Rate limiting storage (in production, use Redis or database)
guest_rate_limits = {}
user_content_rate_limits = {}  # Rate limiting for /api/content endpoint
GUEST_DAILY_LIMIT = 3  # 3 searches per day for guest users
USER_CONTENT_DAILY_LIMIT = 10  # 10 content requests per day for authenticated users
RATE_LIMIT_WINDOW = 24 * 60 * 60  # 24 hours in seconds
MAX_URLS_PER_REQUEST = 5  # Maximum URLs per /api/content request


def check_rate_limit(identifier: str, limit: int, storage: dict) -> bool:
    """Generic rate limiting function."""
    current_time = time.time()

    if identifier not in storage:
        storage[identifier] = {"count": 0, "window_start": current_time}

    rate_data = storage[identifier]

    # Reset window if 24 hours have passed
    if current_time - rate_data["window_start"] > RATE_LIMIT_WINDOW:
        rate_data["count"] = 0
        rate_data["window_start"] = current_time

    # Check if limit exceeded
    if rate_data["count"] >= limit:
        return False

    # Increment count
    rate_data["count"] += 1
    return True


def check_guest_rate_limit(client_ip: str) -> bool:
    """Check if guest user has exceeded rate limit."""
    return check_rate_limit(client_ip, GUEST_DAILY_LIMIT, guest_rate_limits)


def check_user_content_rate_limit(user_id: str) -> bool:
    """Check if authenticated user has exceeded content processing rate limit."""
    return check_rate_limit(user_id, USER_CONTENT_DAILY_LIMIT, user_content_rate_limits)


async def get_current_user_optional(
    authorization: Optional[str] = Header(None),
) -> Optional[str]:
    """Get current user if authenticated, return None if not."""
    if not authorization or not authorization.startswith("Bearer "):
        return None

    token = authorization.split(" ")[1]
    user_id = db_manager.validate_session(token)
    return user_id


# Authentication dependency
async def get_current_user(authorization: Optional[str] = Header(None)) -> str:
    """Extract and validate user from session token"""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Authentication required")

    token = authorization.split(" ")[1]
    user_id = db_manager.validate_session(token)

    if not user_id:
        raise HTTPException(status_code=401, detail="Invalid or expired session")

    return user_id


@app.get("/")
async def root():
    """Basic welcome endpoint"""
    return {"message": "Multimodal Scout API", "version": "1.0.0", "docs": "/docs"}


@app.get("/health")
async def health_check():
    """Health check endpoint for monitoring"""
    try:
        # Test database connection
        with db_manager.get_session() as session:
            session.execute(text("SELECT 1"))
        return {"status": "healthy", "database": "connected"}
    except Exception as e:
        logger.error(f"Health check failed: {e}")
        raise HTTPException(status_code=503, detail="Service unhealthy")


@app.get("/api/config", response_model=ConfigResponse)
async def get_config():
    """Get application configuration values"""
    return ConfigResponse(
        max_urls_per_request=MAX_URLS_PER_REQUEST,
        user_content_daily_limit=USER_CONTENT_DAILY_LIMIT,
        guest_daily_limit=GUEST_DAILY_LIMIT,
    )


@lru_cache(maxsize=1)
def get_cached_topics() -> TopicResponse:
    """Cache static topics data"""
    return TopicResponse(topics=INTERESTED_KEYWORDS)


@app.get("/api/topics", response_model=TopicResponse)
async def get_default_topics():
    """
    Get default interested topics from constants.
    These topics are read-only and cannot be modified by users.
    """
    try:
        logger.info("Fetching default topics from constants")
        response = get_cached_topics()

        # Add cache headers for better client-side caching
        return Response(
            content=response.model_dump_json(),
            media_type="application/json",
            headers={
                "Cache-Control": "public, max-age=3600",  # Cache for 30 minutes
                "Content-Type": "application/json",
            },
        )
    except Exception as e:
        create_user_friendly_error(
            "processing_error", "Unable to load topics at this time.", str(e), 500
        )


@app.post("/api/content/search", response_model=FetchResponse)
async def search_content(
    request: FetchRequest,
    request_obj: Request,
    authorization: Optional[str] = Header(None),
):
    """
    Search for content from various sources based on topics and time range.
    Hybrid access:
    - Authenticated users: Unlimited searches
    - Guest users: 3 searches per day (rate limited by IP)
    """
    try:
        # Check if user is authenticated
        current_user = await get_current_user_optional(authorization)
        client_ip = request_obj.client.host

        if current_user:
            # Authenticated user - unlimited access
            user_type = f"authenticated user {current_user}"
        else:
            # Guest user - check rate limit
            if not check_guest_rate_limit(client_ip):
                remaining_hours = RATE_LIMIT_WINDOW - (
                    time.time() - guest_rate_limits[client_ip]["window_start"]
                )
                raise HTTPException(
                    status_code=429,
                    detail={
                        "error": "rate_limit_exceeded",
                        "message": f"Daily search limit exceeded for guest users ({GUEST_DAILY_LIMIT} searches/day). Please register for unlimited access.",
                        "reset_in_hours": round(remaining_hours / 3600, 1),
                        "current_usage": guest_rate_limits[client_ip]["count"],
                        "daily_limit": GUEST_DAILY_LIMIT,
                    },
                )
            user_type = f"guest user (IP: {client_ip})"

        mode_msg = (
            "discovery mode" if request.discoveryMode else f"topics: {request.topics}"
        )
        logger.info(
            f"{user_type}: Fetching items for {request.selectedDays} days with {mode_msg}"
        )

        pipeline_generator = process_content_pipeline(
            topics=request.topics,
            max_results=request.maxResults,
            research_ratio=request.researchRatio,
            selected_days=request.selectedDays,
            session_id=request.sessionId,
            discovery_mode=request.discoveryMode,
            user_id=current_user,
        )

        final_result = None
        # The pipeline is a generator, so we iterate through it to get the final result.
        # In the non-streaming case, we ignore progress events and just wait for the 'result' event.
        async for event in pipeline_generator:
            if event["type"] == "result":
                final_result = event["data"]
                break

        if final_result:
            logger.info(
                f"Successfully prepared {len(final_result['items'])} items for response"
            )
            return FetchResponse(
                items=final_result["items"],
                total_count=final_result["total_count"],
                sources=final_result["sources"],
            )
        else:
            logger.error("Pipeline did not return a final result.")
            raise HTTPException(
                status_code=500,
                detail="Failed to fetch items: Pipeline finished without result.",
            )

    except Exception as e:
        logger.error(f"Failed to fetch items: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Failed to fetch items: {str(e)}")


@app.post("/api/content/search/stream")
async def search_content_stream(
    request: FetchRequest,
    request_obj: Request,
    authorization: Optional[str] = Header(None),
):
    """
    Search for content with streaming progress updates using the core pipeline.
    Uses Server-Sent Events (SSE) to provide real-time progress.
    Hybrid access:
    - Authenticated users: Unlimited searches
    - Guest users: 3 searches per day (rate limited by IP)
    """

    # Check authentication and rate limiting before starting stream
    current_user = await get_current_user_optional(authorization)
    client_ip = request_obj.client.host

    if current_user:
        user_type = f"authenticated user {current_user}"
    else:
        # Guest user - check rate limit (but don't increment here, increment in stream)
        if client_ip in guest_rate_limits:
            rate_data = guest_rate_limits[client_ip]
            current_time = time.time()

            # Reset window if needed
            if current_time - rate_data["window_start"] > RATE_LIMIT_WINDOW:
                rate_data["count"] = 0
                rate_data["window_start"] = current_time

            if rate_data["count"] >= GUEST_DAILY_LIMIT:
                remaining_hours = RATE_LIMIT_WINDOW - (
                    current_time - rate_data["window_start"]
                )
                raise HTTPException(
                    status_code=429,
                    detail={
                        "error": "rate_limit_exceeded",
                        "message": f"Daily search limit exceeded for guest users ({GUEST_DAILY_LIMIT} searches/day). Please register for unlimited access.",
                        "reset_in_hours": round(remaining_hours / 3600, 1),
                    },
                )
        user_type = f"guest user (IP: {client_ip})"

    async def generate_stream():
        try:
            # Increment rate limit for guest users (after auth check passes)
            if not current_user:
                check_guest_rate_limit(client_ip)

            mode_msg = (
                "discovery mode"
                if request.discoveryMode
                else f"topics: {request.topics}"
            )
            logger.info(
                f"{user_type}: Starting streaming fetch for {request.selectedDays} days with {mode_msg}"
            )

            pipeline_generator = process_content_pipeline(
                topics=request.topics,
                max_results=request.maxResults,
                research_ratio=request.researchRatio,
                selected_days=request.selectedDays,
                session_id=request.sessionId,
                discovery_mode=request.discoveryMode,
                user_id=current_user,
            )

            async for event in pipeline_generator:
                yield f"data: {json.dumps(event)}\n\n"
                # Add a small sleep to allow the client to process the event
                await asyncio.sleep(0.01)

            yield "data: [DONE]\n\n"

        except Exception as e:
            logger.error(f"Failed to fetch items in stream: {e}", exc_info=True)
            error_event = {
                "type": "error",
                "message": f"An unexpected error occurred: {str(e)}",
            }
            yield f"data: {json.dumps(error_event)}\n\n"

    return StreamingResponse(
        generate_stream(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "Access-Control-Allow-Origin": "*",
        },
    )


@app.post("/api/auth/register", response_model=AuthResponse)
async def register_user(request: UserRegistrationRequest):
    """Register a new user"""
    try:
        logger.info(f"Registering new user: {request.email} ({request.username})")
        user_id = db_manager.create_user(
            request.email, request.password, request.username
        )
        session_token = db_manager.create_user_session(user_id)

        return AuthResponse(
            success=True,
            message="User registered successfully",
            session_token=session_token,
            user_id=user_id,
        )
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        create_user_friendly_error(
            "database_error", "Unable to register user. Please try again.", str(e), 500
        )


@app.post("/api/auth/login", response_model=AuthResponse)
async def login_user(request: UserLoginRequest):
    """Login user"""
    try:
        logger.info(f"User login attempt: {request.email}")
        user_id = db_manager.authenticate_user(request.email, request.password)

        if not user_id:
            raise HTTPException(status_code=401, detail="Invalid email or password")

        session_token = db_manager.create_user_session(user_id)

        return AuthResponse(
            success=True,
            message="Login successful",
            session_token=session_token,
            user_id=user_id,
        )
    except HTTPException:
        raise
    except Exception as e:
        create_user_friendly_error(
            "database_error", "Unable to login. Please try again.", str(e), 500
        )


@app.post("/api/auth/logout")
async def logout_user(
    current_user: str = Depends(get_current_user), authorization: str = Header(None)
):
    """Logout user"""
    try:
        token = authorization.split(" ")[1] if authorization else None
        if token:
            db_manager.logout_user(token)
        return {"success": True, "message": "Logged out successfully"}
    except Exception as e:
        logger.error(f"Failed to logout: {e}")
        return {"success": False, "message": "Logout failed"}


@app.get("/api/auth/me", response_model=UserResponse)
async def get_current_user_info(current_user: str = Depends(get_current_user)):
    """Get current user information"""
    try:
        user = db_manager.get_user_by_id(current_user)
        if not user:
            raise HTTPException(status_code=404, detail="User not found")

        return UserResponse(
            user_id=str(user.id),
            email=user.email,
            username=user.username,
            created_at=user.created_at.isoformat(),
            last_login=user.last_login.isoformat() if user.last_login else None,
        )
    except HTTPException:
        raise
    except Exception as e:
        create_user_friendly_error(
            "database_error", "Unable to get user information.", str(e), 500
        )


@app.get("/api/user/preferences", response_model=UserPreferencesResponse)
async def get_user_preferences(current_user: str = Depends(get_current_user)):
    """Get user preferences including custom topics"""
    try:
        custom_topics = db_manager.get_user_custom_topics(current_user)
        return UserPreferencesResponse(custom_topics=custom_topics)
    except Exception as e:
        logger.error(f"Failed to get user preferences: {e}")
        raise HTTPException(status_code=500, detail="Failed to get user preferences")


@app.put("/api/user/preferences")
async def update_user_preferences(
    preferences: UpdateUserPreferencesRequest,
    current_user: str = Depends(get_current_user),
):
    """Update user preferences including custom topics"""
    try:
        success = db_manager.update_user_custom_topics(
            current_user, preferences.custom_topics
        )
        if success:
            return {"success": True, "message": "Preferences updated successfully"}
        else:
            raise HTTPException(status_code=404, detail="User not found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update user preferences: {e}")
        raise HTTPException(status_code=500, detail="Failed to update preferences")


@app.post("/api/bookmarks", response_model=BookmarkResponse)
async def add_bookmark(
    request: BookmarkRequest, current_user: str = Depends(get_current_user)
):
    """Add a bookmark"""
    try:
        logger.info(f"Adding bookmark for user {current_user}: {request.title}")
        bookmark_id = db_manager.add_bookmark(
            user_id=current_user,
            title=request.title,
            link=request.link,
            source_tag=request.source,
            summary=request.summary,
        )
        return BookmarkResponse(
            success=True, message="Bookmark added successfully", bookmark_id=bookmark_id
        )
    except Exception as e:
        create_user_friendly_error(
            "database_error", "Unable to save bookmark. Please try again.", str(e), 500
        )


@app.delete("/api/bookmarks")
async def remove_bookmark(link: str, current_user: str = Depends(get_current_user)):
    """Remove a bookmark by link"""
    try:
        logger.info(f"Removing bookmark for user {current_user}, link: {link}")
        success = db_manager.remove_bookmark(current_user, link)
        if success:
            return BookmarkResponse(
                success=True, message="Bookmark removed successfully"
            )
        else:
            return BookmarkResponse(success=False, message="Bookmark not found")
    except Exception as e:
        logger.error(f"Failed to remove bookmark: {e}")
        raise HTTPException(
            status_code=500, detail=f"Failed to remove bookmark: {str(e)}"
        )


@app.get("/api/bookmarks/check")
async def check_bookmark(link: str, current_user: str = Depends(get_current_user)):
    """Check if a link is bookmarked"""
    try:
        is_bookmarked = db_manager.is_bookmarked(current_user, link)
        return {"is_bookmarked": is_bookmarked}
    except Exception as e:
        logger.error(f"Failed to check bookmark: {e}")
        raise HTTPException(
            status_code=500, detail=f"Failed to check bookmark: {str(e)}"
        )


@app.get("/api/bookmarks/{bookmark_id}")
async def get_bookmark(bookmark_id: str, current_user: str = Depends(get_current_user)):
    """Get a specific bookmark by ID"""
    try:
        bookmark = db_manager.get_bookmark_by_id(current_user, bookmark_id)
        if not bookmark:
            raise HTTPException(status_code=404, detail="Bookmark not found")

        # Get the edited summary if available, otherwise use the original
        summary_edited = getattr(bookmark, "summary_edited", None)
        display_summary = summary_edited or bookmark.summary or "No summary available"

        return {
            "id": str(bookmark.id),
            "title": bookmark.title,
            "link": bookmark.link,
            "summary": display_summary,
            "source": bookmark.source_tag,
            "created_at": bookmark.bookmarked_at.isoformat(),
            "summary_edited": bool(summary_edited),
        }
    except HTTPException:
        raise
    except Exception as e:
        create_user_friendly_error(
            "database_error", "Unable to retrieve bookmark.", str(e), 500
        )


@app.get("/api/bookmarks")
async def get_bookmarks(
    limit: int = 100,
    days: Optional[int] = None,
    current_user: str = Depends(get_current_user),
):
    """Get bookmarks with optional filtering by days back and result limit"""
    try:
        bookmarks = db_manager.get_bookmarks(current_user, limit=limit, days_back=days)
        bookmark_items = []
        for bookmark in bookmarks:
            # Handle both old and new schema gracefully
            summary_edited = getattr(bookmark, "summary_edited", None)
            display_summary = (
                summary_edited or bookmark.summary or "No summary available"
            )
            is_edited = bool(summary_edited)

            # Check for HN comment insights and create combined summary
            comment_insights = None
            comment_count = None
            final_display_summary = display_summary

            if "news.ycombinator.com" in bookmark.link.lower():
                insights_data = db_manager.get_comment_insights(bookmark.link)
                if insights_data:
                    comment_insights = insights_data.insights
                    comment_count = insights_data.comment_count

                    # Create two-section summary for HN bookmarks with comment insights
                    if comment_insights:
                        final_display_summary = f"""**Content Summary:**
{display_summary}

**Community Discussion ({comment_count} comments):**
{comment_insights}"""

            bookmark_items.append(
                ItemResponse(
                    title=bookmark.title,
                    link=bookmark.link,
                    summary=final_display_summary,
                    source=bookmark.source_tag,
                    created_at=bookmark.bookmarked_at.isoformat(),
                    summary_edited=is_edited,
                    comment_insights=comment_insights,
                    comment_count=comment_count,
                )
            )
        return FetchResponse(
            items=bookmark_items, total_count=len(bookmark_items), sources=["Bookmarks"]
        )
    except Exception as e:
        logger.error(f"Failed to get bookmarks: {e}")
        raise HTTPException(
            status_code=500, detail=f"Failed to get bookmarks: {str(e)}"
        )


@app.put("/api/bookmarks/summary")
async def update_bookmark_summary(
    link: str, summary: str, current_user: str = Depends(get_current_user)
):
    """Update a bookmark's summary"""
    try:
        logger.info(f"Updating summary for bookmark: {link}")
        success = db_manager.update_bookmark_summary(current_user, link, summary)
        if success:
            return BookmarkResponse(
                success=True, message="Bookmark summary updated successfully"
            )
        else:
            raise HTTPException(status_code=404, detail="Bookmark not found")
    except Exception as e:
        logger.error(f"Failed to update bookmark summary: {e}")
        raise HTTPException(
            status_code=500, detail=f"Failed to update bookmark summary: {str(e)}"
        )


@app.delete("/api/bookmarks/{bookmark_id}")
async def delete_bookmark(
    bookmark_id: str, current_user: str = Depends(get_current_user)
):
    """Delete a specific bookmark by ID"""
    try:
        logger.info(f"Removing bookmark with ID: {bookmark_id}")
        success = db_manager.remove_bookmark_by_id(current_user, bookmark_id)
        if success:
            return {"message": "Bookmark deleted successfully"}
        else:
            raise HTTPException(status_code=404, detail="Bookmark not found")
    except HTTPException:
        raise
    except Exception as e:
        create_user_friendly_error(
            "database_error", "Unable to delete bookmark.", str(e), 500
        )


@app.patch("/api/bookmarks/{bookmark_id}")
async def update_bookmark(
    bookmark_id: str, request: dict, current_user: str = Depends(get_current_user)
):
    """Update a bookmark's summary by ID"""
    try:
        summary = request.get("summary", "")
        logger.info(f"Updating summary for bookmark: {bookmark_id}")
        success = db_manager.update_bookmark_summary_by_id(
            current_user, bookmark_id, summary
        )
        if success:
            return {"message": "Bookmark updated successfully"}
        else:
            raise HTTPException(status_code=404, detail="Bookmark not found")
    except HTTPException:
        raise
    except Exception as e:
        create_user_friendly_error(
            "database_error", "Unable to update bookmark.", str(e), 500
        )


@app.post("/api/content", response_model=UploadLinkResponse)
async def create_content(
    request: UploadLinkRequest, current_user: str = Depends(get_current_user)
):
    """Create content items from user-provided links with security restrictions"""
    try:
        # Security check 1: URL count validation (max 5 URLs per request)
        if len(request.urls) > MAX_URLS_PER_REQUEST:
            raise HTTPException(
                status_code=400,
                detail={
                    "error": "validation_error",
                    "message": f"Maximum {MAX_URLS_PER_REQUEST} URLs allowed per request",
                    "provided": len(request.urls),
                    "limit": MAX_URLS_PER_REQUEST,
                },
            )

        # Security check 2: Rate limiting (10 requests per 24h for authenticated users)
        if not check_user_content_rate_limit(current_user):
            remaining_hours = RATE_LIMIT_WINDOW - (
                time.time() - user_content_rate_limits[current_user]["window_start"]
            )
            raise HTTPException(
                status_code=429,
                detail={
                    "error": "rate_limit_exceeded",
                    "message": f"Daily content processing limit exceeded ({USER_CONTENT_DAILY_LIMIT} requests/day)",
                    "reset_in_hours": round(remaining_hours / 3600, 1),
                    "current_usage": user_content_rate_limits[current_user]["count"],
                    "daily_limit": USER_CONTENT_DAILY_LIMIT,
                },
            )

        logger.info(
            f"Processing {len(request.urls)} uploaded links for user {current_user}"
        )

        results = []
        failed_urls = []
        success_count = 0

        for url_obj in request.urls:
            try:
                url = str(url_obj)
                logger.info(f"Processing URL: {url}")

                # Check if already bookmarked
                if db_manager.is_bookmarked(current_user, url):
                    failed_urls.append(url)
                    logger.info(f"URL already bookmarked, skipping: {url}")
                    continue

                # Extract title
                title = extract_title_from_url(url_obj)
                if not title:
                    title = f"Article from {url_obj.host}"

                # Extract content for categorization
                article_text = _fetch_article_text(url_obj)
                if not article_text:
                    article_text = ""

                # Generate summary
                summary = generate_summary_from_link(url_obj, title)

                # Categorize content
                source_tag = categorize_content(title, article_text, url)

                # Add to bookmarks
                bookmark_id = db_manager.add_bookmark(
                    user_id=current_user,
                    title=title,
                    link=url,
                    source_tag=source_tag,
                    summary=summary,
                )

                # Also cache the summary for future reference
                db_manager.add_summary(url, summary)

                results.append(
                    {
                        "url": url,
                        "title": title,
                        "summary": summary,
                        "source_tag": source_tag,
                        "bookmark_id": bookmark_id,
                    }
                )
                success_count += 1

                logger.info(f"Successfully processed: {title} (Category: {source_tag})")

            except Exception as url_error:
                logger.error(f"Failed to process URL {url}: {url_error}")
                failed_urls.append(str(url_obj))

        # Determine response
        if success_count == 0:
            return UploadLinkResponse(
                success=False,
                message="Failed to process any URLs",
                results=[],
                failed_urls=failed_urls,
            )
        elif len(failed_urls) > 0:
            return UploadLinkResponse(
                success=True,
                message=f"Processed {success_count} URLs successfully, {len(failed_urls)} failed",
                results=results,
                failed_urls=failed_urls,
            )
        else:
            return UploadLinkResponse(
                success=True,
                message=f"Successfully processed all {success_count} URLs",
                results=results,
                failed_urls=[],
            )

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to process uploaded links: {e}", exc_info=True)
        raise HTTPException(
            status_code=500, detail=f"Failed to process links: {str(e)}"
        )


@app.get("/api/bookmarks/export")
async def export_bookmarks(current_user: str = Depends(get_current_user)):
    """Export all bookmarks to Excel file"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from io import BytesIO
        from datetime import datetime

        logger.info("Starting bookmark export to Excel")

        # Get all bookmarks
        bookmarks = db_manager.get_bookmarks(current_user)

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Bookmarks"

        # Define headers
        headers = ["Title", "Summary", "Source URL", "Source Date"]

        # Add headers with styling
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Add bookmark data
        for row, bookmark in enumerate(bookmarks, 2):
            # Handle both edited and original summaries
            summary_edited = getattr(bookmark, "summary_edited", None)
            display_summary = (
                summary_edited or bookmark.summary or "No summary available"
            )

            # Format date
            source_date = (
                bookmark.bookmarked_at.strftime("%Y-%m-%d %H:%M:%S")
                if bookmark.bookmarked_at
                else "Unknown"
            )

            ws.cell(row=row, column=1, value=bookmark.title)
            ws.cell(row=row, column=2, value=display_summary)
            ws.cell(row=row, column=3, value=bookmark.link)
            ws.cell(row=row, column=4, value=source_date)

        # Adjust column widths
        ws.column_dimensions["A"].width = 50  # Title
        ws.column_dimensions["B"].width = 80  # Summary
        ws.column_dimensions["C"].width = 60  # URL
        ws.column_dimensions["D"].width = 20  # Date

        # Save to BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # Generate filename with current timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"multimodal_scout_bookmarks_{timestamp}.xlsx"

        logger.info(f"Successfully exported {len(bookmarks)} bookmarks to Excel")

        # Return Excel file as response
        return Response(
            content=excel_buffer.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename={filename}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            },
        )

    except Exception as e:
        logger.error(f"Failed to export bookmarks: {e}", exc_info=True)
        raise HTTPException(
            status_code=500, detail=f"Failed to export bookmarks: {str(e)}"
        )


@app.post("/pipeline")
async def pipeline_cron_job(authorization: Optional[str] = Header(None)):
    """
    Pipeline endpoint for scheduled jobs.
    Cloud: Requires pipeline-secret from Secret Manager
    Local: No authentication required for development
    """
    from .config import config

    # In cloud environment, require pipeline secret
    if config.is_cloud_environment:
        pipeline_secret = config.get_secret("pipeline-secret")
        if not pipeline_secret:
            logger.error("Pipeline endpoint disabled - pipeline-secret not configured")
            raise HTTPException(
                status_code=503, detail="Pipeline endpoint not configured"
            )

        # Verify authorization header
        if not authorization or not authorization.startswith("Bearer "):
            logger.warning("Pipeline endpoint accessed without valid authorization")
            raise HTTPException(status_code=401, detail="Authorization required")

        # Extract and validate token
        token = authorization.split(" ", 1)[1] if " " in authorization else ""
        if token != pipeline_secret:
            logger.warning("Pipeline endpoint accessed with invalid secret")
            raise HTTPException(status_code=403, detail="Access denied")

        logger.info("🔓 Pipeline authenticated via Secret Manager")
    else:
        # Local development - no authentication required
        logger.info("🔓 Pipeline accessed in local development mode")
    try:
        logger.info("🚀 Starting pipeline via HTTP endpoint...")

        # Run pipeline with Cloud Scheduler (every 30 minutes)
        pipeline_generator = process_content_pipeline(
            topics=[],  # No topic filtering for cron jobs
            max_results=50,  # Match the config file setting
            research_ratio=0.5,
            selected_days=1,
        )

        # Process pipeline events (simplified for HTTP context)
        final_result = None
        async for event in pipeline_generator:
            if event.get("type") == "result":
                final_result = event.get("data", {})
                break

        if final_result:
            total_items = final_result.get("total_count", 0)
            sources = final_result.get("sources", [])
            logger.info(
                f"✅ Pipeline completed: {total_items} items from {len(sources)} sources"
            )

            return {
                "status": "success",
                "message": "Pipeline completed successfully",
                "total_items": total_items,
                "sources": sources,
            }
        else:
            logger.warning("⚠️ Pipeline completed without results")
            return {
                "status": "success",
                "message": "Pipeline completed with no new results",
                "total_items": 0,
                "sources": [],
            }

    except Exception as e:
        logger.error(f"❌ Pipeline failed: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail={"status": "error", "message": f"Pipeline failed: {str(e)}"},
        )


@app.get("/api/bookmarks/export/chrome")
async def export_chrome_bookmarks(
    selected_tags: str = None,
    search_query: str = None,
    export_format: str = "html",
    current_user: str = Depends(get_current_user),
):
    """Export bookmarks in Chrome-compatible HTML format with optional filtering"""
    try:
        from datetime import datetime
        from html import escape
        from io import StringIO

        logger.info(f"Starting bookmark export in {export_format} format")

        # Get all bookmarks
        bookmarks = db_manager.get_bookmarks(current_user)

        # Apply filters if provided
        if selected_tags or search_query:
            filtered_bookmarks = []
            selected_tag_set = set()

            if selected_tags:
                # Parse comma-separated tags
                selected_tag_set = set(
                    tag.strip() for tag in selected_tags.split(",") if tag.strip()
                )

            for bookmark in bookmarks:
                # Text search filter
                if search_query and search_query.strip():
                    query = search_query.lower()
                    title_match = bookmark.title and query in bookmark.title.lower()
                    summary_match = (
                        bookmark.summary and query in bookmark.summary.lower()
                    )
                    if not (title_match or summary_match):
                        continue

                # Tag filter
                if selected_tag_set:
                    tag_matches = False
                    # Check source tag
                    if bookmark.source_tag in selected_tag_set:
                        tag_matches = True
                    # TODO: Check matched keywords if available in bookmark model

                    if not tag_matches:
                        continue

                filtered_bookmarks.append(bookmark)

            bookmarks = filtered_bookmarks
            logger.info(f"Filtered bookmarks: {len(bookmarks)} items")

        if export_format == "markdown":
            # Generate Markdown content
            markdown_buffer = StringIO()
            markdown_buffer.write("# Bookmarks\n\n")

            for bookmark in bookmarks:
                title = (
                    (bookmark.title or "Untitled")
                    .replace("\n", " ")
                    .replace("\r", " ")
                    .strip()
                )
                summary = (
                    (bookmark.summary or "No description available.")
                    .replace("\n", " ")
                    .replace("\r", " ")
                    .strip()
                )
                markdown_buffer.write(f"## {title}\n\n")
                markdown_buffer.write(f"{summary}\n\n")

            # Generate filename with timestamp
            timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"multimodal_scout_bookmarks_{timestamp_str}.md"

            logger.info(f"Successfully exported {len(bookmarks)} bookmarks to Markdown")

            markdown_content = markdown_buffer.getvalue()
            markdown_buffer.close()

            return Response(
                content=markdown_content.encode("utf-8"),
                media_type="text/markdown",
                headers={"Content-Disposition": f"attachment; filename={filename}"},
            )

        # Group bookmarks by source for subfolder organization
        research_bookmarks = []
        industry_bookmarks = []

        for bookmark in bookmarks:
            source = bookmark.source_tag.lower() if bookmark.source_tag else ""
            if source == "research":
                research_bookmarks.append(bookmark)
            else:
                industry_bookmarks.append(bookmark)

        # Generate Chrome bookmark HTML format
        timestamp = int(datetime.now().timestamp())

        html_buffer = StringIO()
        html_buffer.write(
            f"""<!DOCTYPE NETSCAPE-Bookmark-file-1>
<!-- This is an automatically generated file.
     It will be read and overwritten.
     DO NOT EDIT! -->
<META HTTP-EQUIV="Content-Type" CONTENT="text/html; charset=UTF-8">
<TITLE>Bookmarks</TITLE>
<H1>Bookmarks</H1>
<DL><p>
    <DT><H3 ADD_DATE="{timestamp}" LAST_MODIFIED="{timestamp}" PERSONAL_TOOLBAR_FOLDER="true">Multimodal Scout</H3>
    <DL><p>
        <DT><H3 ADD_DATE="{timestamp}" LAST_MODIFIED="{timestamp}">Research</H3>
        <DL><p>
"""
        )

        # Add research bookmarks
        for bookmark in research_bookmarks:
            add_date = (
                int(bookmark.bookmarked_at.timestamp())
                if bookmark.bookmarked_at
                else timestamp
            )
            title = escape(
                (bookmark.title or "Untitled")
                .replace("\n", " ")
                .replace("\r", " ")
                .strip()
            )
            url = escape(bookmark.link)
            summary = escape(
                (bookmark.summary or "").replace("\n", " ").replace("\r", " ").strip()
            )
            html_buffer.write(
                f'            <DT><A HREF="{url}" ADD_DATE="{add_date}" DESCRIPTION="{summary}">{title}</A>\n'
            )

        html_buffer.write(
            f"""        </DL><p>
        <DT><H3 ADD_DATE="{timestamp}" LAST_MODIFIED="{timestamp}">Industry</H3>
        <DL><p>
"""
        )

        # Add industry bookmarks
        for bookmark in industry_bookmarks:
            add_date = (
                int(bookmark.bookmarked_at.timestamp())
                if bookmark.bookmarked_at
                else timestamp
            )
            title = escape(
                (bookmark.title or "Untitled")
                .replace("\n", " ")
                .replace("\r", " ")
                .strip()
            )
            url = escape(bookmark.link)
            summary = escape(
                (bookmark.summary or "").replace("\n", " ").replace("\r", " ").strip()
            )
            html_buffer.write(
                f'            <DT><A HREF="{url}" ADD_DATE="{add_date}" DESCRIPTION="{summary}">{title}</A>\n'
            )

        html_buffer.write(
            """        </DL><p>
    </DL><p>
</DL><p>
"""
        )

        # Generate filename with timestamp
        timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"multimodal_scout_chrome_bookmarks_{timestamp_str}.html"

        logger.info(
            f"Successfully exported {len(bookmarks)} bookmarks for Chrome ({len(research_bookmarks)} research, {len(industry_bookmarks)} industry)"
        )

        html_content = html_buffer.getvalue()
        html_buffer.close()

        return Response(
            content=html_content.encode("utf-8"),
            media_type="text/html",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        logger.error(f"Failed to export bookmarks: {e}", exc_info=True)
        raise HTTPException(
            status_code=500, detail=f"Failed to export bookmarks: {str(e)}"
        )


if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=8000, log_level="info")
